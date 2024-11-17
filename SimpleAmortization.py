import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def generate_amortization_schedule(loan_amount, annual_interest_rate, payment, prepayment_func, months, start_date):
    """Generates an amortization schedule as a list of dictionaries."""
    from datetime import datetime, timedelta

    schedule = []
    opening_balance = loan_amount
    monthly_interest_rate = annual_interest_rate / 12 / 100
    current_date = datetime.strptime(start_date, "%m/%d/%Y")
    
    for period in range(months + 1):
        if period == 0:  # Initial period
            schedule.append({
                "Period": period,
                "Date": current_date.strftime("%m/%d/%Y"),
                "Opening Balance": 0,
                "Payment": "-",
                "Prepayment": "-",
                "Interest": "-",
                "Principal": 0,
                "Closing Balance": loan_amount,
            })
            current_date += timedelta(days=30)  # Approximate one month
            continue
        
        prepayment = prepayment_func(period) if callable(prepayment_func) else 0
        interest = round(opening_balance * monthly_interest_rate, 2)
        principal = round((payment + prepayment) - interest, 2)
        closing_balance = round(opening_balance - principal, 2)
        
        if period == months or closing_balance < 0:
            principal += closing_balance
            closing_balance = 0
        
        schedule.append({
            "Period": period,
            "Date": current_date.strftime("%m/%d/%Y"),
            "Opening Balance": opening_balance,
            "Payment": payment,
            "Prepayment": prepayment,
            "Interest": interest,
            "Principal": principal,
            "Closing Balance": closing_balance,
        })
        
        opening_balance = closing_balance
        current_date += timedelta(days=30)  # Approximate one month
        
        if closing_balance == 0:
            break

    return pd.DataFrame(schedule)

def create_excel_with_loans(output_file, loans_data):
    """
    Creates an Excel file with amortization schedules for multiple loans.

    :param output_file: Path to the output Excel file.
    :param loans_data: List of dictionaries containing loan details.
                       Each dictionary should have `loan_number`, `loan_amount`, 
                       `annual_interest_rate`, `payment`, `prepayment_func`, `months`, `start_date`.
    """
    # Create a workbook
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove the default blank sheet

    for loan in loans_data:
        # Generate the amortization schedule for the loan
        schedule_df = generate_amortization_schedule(
            loan_amount=loan["loan_amount"],
            annual_interest_rate=loan["annual_interest_rate"],
            payment=loan["payment"],
            prepayment_func=loan["prepayment_func"],
            months=loan["months"],
            start_date=loan["start_date"]
        )

        # Add a new worksheet with the loan number as the sheet name
        sheet_name = f"Loan_{loan['loan_number']}"
        worksheet = workbook.create_sheet(title=sheet_name)

        # Write DataFrame to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(schedule_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the workbook to the specified output file
    workbook.save(output_file)
    print(f"Excel file with loan schedules saved as '{output_file}'")

# Example Usage
def example_prepayment_func(period):
    return round(150 - (period * 0.5), 2) if period <= 36 else 0

loans = [
    {
        "loan_number": "001",
        "loan_amount": 35000,
        "annual_interest_rate": 8,
        "payment": 1096.77,
        "prepayment_func": example_prepayment_func,
        "months": 36,
        "start_date": "09/01/2023"
    },
    {
        "loan_number": "002",
        "loan_amount": 45000,
        "annual_interest_rate": 10,
        "payment": 1400.50,
        "prepayment_func": lambda period: 100 if period % 2 == 0 else 50,
        "months": 24,
        "start_date": "10/01/2023"
    }
]

# Generate Excel file
output_file_path = "loan_schedules.xlsx"
create_excel_with_loans(output_file_path, loans)



